"""
Created By : Hamza Malik
Dated   : 09/21/2020
BOT TITLE : Translator

"""

from selenium import webdriver
from openpyxl import *
import time
import glob
import os
from pathlib import Path
import re

#-------------------------------------------------------
#main folder


foldername='a'
desk=os.path.expanduser(r"~\Desktop")
my_directory=r'{}\{}'.format(desk,foldername)
"""
Below will get subfolders list from my directory
"""
#folders
folders_list = subdirs = [os.path.join(my_directory, o) for o in os.listdir(my_directory) if
                          os.path.isdir(os.path.join(my_directory,o))]
print("Sub-Folders List: ", folders_list)
excel_files_list=[]
for folder in folders_list:
    extension = 'xlsx'
    os.chdir(folder)
    result = glob.glob('*.{}'.format(extension))
    print(result)
    for i in result:
        excel_files_list.append('{}\{}'.format(folder,i))

print("All Excel Files : ", excel_files_list)

#-------------------------------------------------------

website_url="https://papago.naver.com/"


driver = webdriver.Chrome(r'C:\Users\dell\Desktop\George\chromedriver') 

def main(column_list,file_location):
    driver.get(website_url)
    for index in range (0, len(column_list)):
        #input korean text
        time.sleep(4)
        try:
            input_text = driver.find_element_by_xpath('''//textarea[@rows='1']''')
            input_text.clear()
        except Exception as e:
            print(e)
        finally:
            time.sleep(5)
            input_text = driver.find_element_by_xpath('''//textarea[@rows='1']''')
            input_text.clear()


        input_text.send_keys(column_list[index])
        time.sleep(1)

        try:
            translate_button = driver.find_element_by_xpath('''//div[@class='btn_translation___b0nPG']//button[1]''')
        except Exception as e:
            print(e)
        finally:
            time.sleep(5)
            translate_button = driver.find_element_by_xpath('''//div[@class='btn_translation___b0nPG']//button[1]''')

        translate_button.click()

        #gettting translation
        time.sleep(2)
        try:
            translation_text= driver.find_element_by_xpath('''//div[@id='targetEditArea']//div[1]''').text
        except Exception as e:
            print(e)
        finally:
            time.sleep(5)
            translation_text = driver.find_element_by_xpath('''//div[@id='targetEditArea']//div[1]''').text
            
        print("Translation: ", translation_text)
        #driver.implicitly_wait(200)
        time.sleep(1)
        input_text.clear()
        #puting translation excel file
        wb=load_workbook(file_location)
        ws=wb["your sheet"]
        wcell1=ws.cell(index+1,7)
        wcell1.value=translation_text
        wb.save(file_location)
    #close driver
    driver.close()






if __name__ == "__main__":
    # main loop which does everything
    for file in excel_files_list:
        current_file_location = file

        wb = load_workbook(current_file_location)  # Work Book
        ws = wb.worksheets[0]  # Work Sheet
        column = ws['F']  # Column
        column_list = [column[x].value for x in range(len(column))]
        print(type(column_list), column_list)
        main(column_list, current_file_location)
    # --------------------------------------------------------------------



