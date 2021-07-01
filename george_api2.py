
"""
Created By : Hamza Malik
Dated   : 09/21/2020
BOT TITLE : Translator

"""


from openpyxl import *
import glob
import json
import os
import urllib.request


#--------------------

"""
            Please Exchange the api with your accounts

"""



client_id = "tUWw7wjhM_wCAHnF5bNZ" # 개발자센터에서 발급받은 Client ID 값
client_secret = "zgH9af3SBp" # 개발자센터에서 발급받은 Client Secret 값



# -------------------------------------------------------
# main folder

foldername='a'
desk=os.path.expanduser(r"~\Desktop")
my_directory=r'{}\{}'.format(desk,foldername)

"""
Below will get subfolders list from my directory
"""
# folders
folders_list = subdirs = [os.path.join(my_directory, o) for o in os.listdir(my_directory) if
                          os.path.isdir(os.path.join(my_directory, o))]
print("Sub-Folders List: ", folders_list)
excel_files_list = []
for folder in folders_list:
    extension = 'xlsx'
    os.chdir(folder)
    result = glob.glob('*.{}'.format(extension))
    print(result)
    for i in result:
        excel_files_list.append('{}\{}'.format(folder, i))

print("All Excel Files : ", excel_files_list)

# -------------------------------------------------------



def main(column_list, file_location):
    for index in range(0, len(column_list)):
        # input korean text
        encText = urllib.parse.quote(column_list[index])
        data = "source=ko&target=en&text=" + encText
        url = "https://openapi.naver.com/v1/papago/n2mt"
        request = urllib.request.Request(url)
        request.add_header("X-Naver-Client-Id", client_id)
        request.add_header("X-Naver-Client-Secret", client_secret)
        response = urllib.request.urlopen(request, data=data.encode("utf-8"))
        rescode = response.getcode()
        if (rescode == 200):
            response_body = response.read()
            dic = json.loads(response_body)
            text_get = dic["message"]["result"]
            translation_text=text_get["translatedText"]

        else:
            print("Error Code:" + rescode)
            continue

        print("Translation: ", translation_text)
        # driver.implicitly_wait(200)
        # puting translation excel file
        wb = load_workbook(file_location)
        ws = wb.worksheets[0]
        wcell1 = ws.cell(index + 1, 7)
        wcell1.value = translation_text
        wb.save(file_location)



if __name__ == "__main__":
    # main loop which does everything
    for file in excel_files_list:
        current_file_location = file

        wb = load_workbook(current_file_location)  # Work Book
        ws = wb.worksheets[0]  # Work Sheet
        column = ws['F']  # Column
        column_list = [column[x].value for x in range(len(column))]
        print(type(column_list), column_list)
        main(column_list, current_file_location)
    # --------------------------------------------------------------------



